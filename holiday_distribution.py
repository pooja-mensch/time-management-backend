from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from werkzeug.utils import secure_filename
import tempfile
import uuid
import shutil


import openpyxl
import os
import pandas as pd
import warnings
from openpyxl import load_workbook
import random
from openpyxl.utils import coordinate_to_tuple
import re

class HolidayTool:
    
    def __init__(self, path):
        self.path = path
        self.emp_list = []
        self.meta_data = None
        self.max_emp = 30
        
        self.bl_mapping = {
            2020: {
                'Baden-Württemberg': 4, 'Bayern (katholisch)': 5, 'Bayern': 6,
                'Berlin': 7, 'Brandenburg': 8, 'Bremen': 9, 'Hamburg': 10,
                'Hessen': 11, 'Mecklenburg-Vorpommern': 12, 'Niedersachsen': 13,
                'Nordrhein-Westfahlen': 14, 'Rheinland-Pfalz': 15, 'Saarland': 16,
                'Sachsen': 17, 'Sachsen-Anhalt': 18, 'Schleswig-Holstein': 19,
                'Thüringen': 20
            },
            2021: {
                'Baden-Württemberg': 25, 'Bayern (katholisch)': 26, 'Bayern': 27,
                'Berlin': 28, 'Brandenburg': 29, 'Bremen': 30, 'Hamburg': 31,
                'Hessen': 32, 'Mecklenburg-Vorpommern': 33, 'Niedersachsen': 34,
                'Nordrhein-Westfahlen': 35, 'Rheinland-Pfalz': 36, 'Saarland': 37,
                'Sachsen': 38, 'Sachsen-Anhalt': 39, 'Schleswig-Holstein': 40,
                'Thüringen': 41
            },
            2022: {
                'Baden-Württemberg': 46, 'Bayern (katholisch)': 47, 'Bayern': 48,
                'Berlin': 49, 'Brandenburg': 50, 'Bremen': 51, 'Hamburg': 52,
                'Hessen': 53, 'Mecklenburg-Vorpommern': 54, 'Niedersachsen': 55,
                'Nordrhein-Westfahlen': 56, 'Rheinland-Pfalz': 57, 'Saarland': 58,
                'Sachsen': 59, 'Sachsen-Anhalt': 60, 'Schleswig-Holstein': 61,
                'Thüringen': 62
            },
             2023: {
                'Baden-Württemberg': 67, 'Bayern (katholisch)': 68, 'Bayern': 69,
                'Berlin': 70, 'Brandenburg': 71, 'Bremen': 72, 'Hamburg': 73,
                'Hessen': 74, 'Mecklenburg-Vorpommern': 75, 'Niedersachsen': 76,
                'Nordrhein-Westfahlen': 77, 'Rheinland-Pfalz': 78, 'Saarland': 79,
                'Sachsen': 80, 'Sachsen-Anhalt': 81, 'Schleswig-Holstein': 82,
                'Thüringen': 83
            },
            2024: {
                'Baden-Württemberg': 88 , 'Bayern (katholisch)': 89, 'Bayern': 90,
                'Berlin': 91, 'Brandenburg': 92, 'Bremen': 93, 'Hamburg': 94,
                'Hessen': 95, 'Mecklenburg-Vorpommern': 96, 'Niedersachsen': 97,
                'Nordrhein-Westfahlen': 98, 'Rheinland-Pfalz': 99, 'Saarland': 100,
                'Sachsen': 101, 'Sachsen-Anhalt': 102, 'Schleswig-Holstein': 103,
                'Thüringen': 104
            },
            2025: {
                'Baden-Württemberg': 109 , 'Bayern (katholisch)': 110, 'Bayern': 111,
                'Berlin': 112, 'Brandenburg': 113, 'Bremen': 114, 'Hamburg': 115,
                'Hessen': 116, 'Mecklenburg-Vorpommern': 117, 'Niedersachsen': 118,
                'Nordrhein-Westfahlen': 119, 'Rheinland-Pfalz': 120, 'Saarland': 121,
                'Sachsen': 122, 'Sachsen-Anhalt': 123, 'Schleswig-Holstein': 124,
                'Thüringen': 125
            },

            2026: {
                'Baden-Württemberg': 130 , 'Bayern (katholisch)': 131, 'Bayern': 132,
                'Berlin': 133, 'Brandenburg': 134, 'Bremen': 135, 'Hamburg': 136,
                'Hessen': 137, 'Mecklenburg-Vorpommern': 138, 'Niedersachsen': 139,
                'Nordrhein-Westfahlen': 140, 'Rheinland-Pfalz': 141, 'Saarland': 142,
                'Sachsen': 143, 'Sachsen-Anhalt': 144, 'Schleswig-Holstein': 145,
                'Thüringen': 146
            },

            2027: {
                'Baden-Württemberg': 151 , 'Bayern (katholisch)': 152, 'Bayern': 153,
                'Berlin': 154, 'Brandenburg': 155, 'Bremen': 156, 'Hamburg': 157,
                'Hessen': 158, 'Mecklenburg-Vorpommern': 159, 'Niedersachsen': 160,
                'Nordrhein-Westfahlen': 161, 'Rheinland-Pfalz': 162, 'Saarland': 163,
                'Sachsen': 164, 'Sachsen-Anhalt': 165, 'Schleswig-Holstein': 166,
                'Thüringen': 167
            }
        }
        self.file_format = None  # 10, 20, or 30
        self.actual_years = []   
        self.yr_offset = {}
        
        self.base_row = 3
        self.fei_first = 4
        self.fei_last = 377
        self.ist_first = 15
        self.ist_last = 390
        self.emp_start_row = 6
        
        warnings.filterwarnings("ignore", message="Conditional Formatting extension is not supported and will be removed")
    
    def detect_file_format_and_years(self):
        """Detect file format (10/20/30 employees) and read actual years from the file"""
        try:
            wb = load_workbook(self.path, data_only=True)
            
            sheet = wb['IST Stunden']
            
            # First, determine the step size by checking common patterns
            patterns_to_test = [
                {'format': 10, 'step': 16},  # 10 employees: B1, B17, B33, B49, etc.
                {'format': 20, 'step': 26},  # 20 employees: B1, B27, B53, B79, etc.
                {'format': 30, 'step': 36}   # 30 employees: B1, B37, B73, B109, etc.
            ]
            
            detected_format = None
            step_size = None
            
            # Test each pattern to see which one has valid years
            for pattern in patterns_to_test:
                test_positions = [3, 3 + pattern['step'], 3 + 2*pattern['step']]
                valid_years = 0
                
                for pos in test_positions:
                    try:
                        cell_value = sheet.cell(row=pos, column=2).value
                        if cell_value and str(cell_value).isdigit():
                            year = int(cell_value)
                            if 2015 <= year <= 2030:  # Reasonable year range
                                valid_years += 1
                    except:
                        pass
                
                if valid_years >= 2:  # At least 2 valid years found
                    detected_format = pattern['format']
                    step_size = pattern['step']
                    break
            
            if not detected_format:
                raise Exception("Could not detect file format")
            
            # Now scan ALL positions to find all years
            all_years = []
            position = 3 # Start at B1
            max_positions_to_check = 23  # Safety limit
            
            for i in range(max_positions_to_check):
                current_position = 3 + (i * step_size)
                
                try:
                    cell_value = sheet.cell(row=current_position, column=2).value
                    if cell_value and str(cell_value).isdigit():
                        year = int(cell_value)
                        if 2015 <= year <= 2030:  # Reasonable year range
                            all_years.append((year, current_position))
                            #print(f"Found year {year} at position B{current_position}")
                        else:
                            # If we hit an invalid year, we might have reached the end
                            break
                    else:
                        # If we hit an empty cell, we've probably reached the end
                        break
                except:
                    # If we can't read the cell, we've probably reached the end
                    break
            
            if len(all_years) < 2:
                raise Exception("Not enough valid years found")
            
            # Verify years are consecutive
            years_only = [y[0] for y in all_years]
            years_only.sort()
            
            # Check if years are consecutive (allowing for some gaps)
            if not all(years_only[i+1] - years_only[i] <= 2 for i in range(len(years_only)-1)):
                print("Warning: Years might not be consecutive")
            
            self.file_format = detected_format
            self.actual_years = years_only
            
            # Create year offset mapping using actual positions found
            self.yr_offset = {}
            for year, position in all_years:
                self.yr_offset[year] = position
            
            #print(f"Detected format: {detected_format} employees")
            #print(f"Years found: {self.actual_years}")
            #print(f"Year offsets: {self.yr_offset}")
            
            # Verify we have holiday data for all detected years
            missing_years = []
            for year in self.actual_years:
                if year not in self.bl_mapping:
                    missing_years.append(year)
            
            if missing_years:
                #print(f"Warning: No holiday data available for years: {missing_years}")
                # Remove years without holiday data
                self.actual_years = [y for y in self.actual_years if y in self.bl_mapping]
                #print(f"Will process only years with holiday data: {self.actual_years}")
            
            return True
            
        except Exception as e:
            #print(f"Error detecting file format: {e}")
            return False
    
    def get_metadata(self):
        try:
            self.meta_data = pd.read_excel(self.path, sheet_name="MA Übersicht",skiprows=2)
            return True
        except Exception as e:
            raise Exception(f"Metadata error: {e}")
    
    def find_employees(self):
        if self.meta_data is None:
             raise Exception("No metadata available")
        count = 0
        for i, row in self.meta_data.iterrows():
            if count >= self.max_emp:
          
                break

            if pd.isna(row.get("Vorname", None)) or pd.isna(row.get("Nachname", None)) or pd.isna(row.get("Bundesland", None)):
                continue

            start_dt = row.iloc[19]
            end_dt = row.iloc[20]

            emp_info = {
                'full_name': f"{row['Vorname']} {row['Nachname']}",
                'state': row["Bundesland"],
                'orig_idx': i,
                'emp_num': count + 1,
                'row_offset': count,
                'start': start_dt,
                'end': end_dt
            }

            self.emp_list.append(emp_info)
            count += 1

           
    
    def process_employee_holidays(self, emp_data, wb,formulas_comments):
        sheet = wb['IST Stunden']

        try:
            fei_file_path = os.path.join(os.path.dirname(__file__), "Feiertage.xlsx")
            fei_wb = load_workbook(fei_file_path, data_only=True)
            fei_sheet = fei_wb.active
        except Exception as e:
            raise Exception(f"Can't load holidays file: {e}")
           

        #print(f"\nDoing {emp_data['full_name']} in {emp_data['state']}...")

        emp_row_pos = self.emp_start_row + emp_data['row_offset']

        start_dt = pd.to_datetime(emp_data.get('start'))
        end_dt = pd.to_datetime(emp_data.get('end'))

        if pd.isna(start_dt) or pd.isna(end_dt):
            #print(f"  Missing dates for {emp_data['full_name']}")
            return False

        #print(f"  Working: {start_dt} to {end_dt}")

        total_marked = 0
        total_skipped = 0

        try:
            wb_data_only = load_workbook(self.path, data_only=True)
            sheet_data_only = wb_data_only['IST Stunden']
        except Exception as e:
            #print(f"Warning: Could not load workbook with data_only=True: {e}")
            sheet_data_only = None
        for yr in self.actual_years:
            if yr not in self.bl_mapping:
                #print(f"  No holiday data for year {yr}")
                continue
            if emp_data['state'] not in self.bl_mapping[yr]:
                #print(f"  Can't find {emp_data['state']} for {yr}")
                continue

            src_row = self.bl_mapping[yr][emp_data['state']]
            holiday_data = []
            
            for col in range(self.fei_first, self.fei_last + 1):
                val = fei_sheet.cell(row=src_row, column=col).value
                holiday_data.append(val)

            processed_holidays = []
            for v in holiday_data:
                if v is not None and str(v).strip() != "":
                    processed_holidays.append('f')
                else:
                    processed_holidays.append(None)
            print(f"  Holiday data for {yr}: {processed_holidays[:10]}... (first 10 of {len(processed_holidays)})")

            target_start = self.yr_offset[yr] + 3
            target_row_num = target_start + emp_data['row_offset']
            date_row = target_start - 2

            marked_this_yr = 0
            skipped_this_yr = 0

            for idx in range(len(processed_holidays)):
                if idx >= len(processed_holidays):
                    break
                    
                col_num = self.ist_first + idx
                if col_num > self.ist_last:
                    break

                if processed_holidays[idx] == 'f':
                    dt_cell = sheet_data_only.cell(row=date_row, column=col_num).value if sheet_data_only else sheet.cell(row=date_row, column=col_num).value
                    dt_obj = pd.to_datetime(dt_cell, errors='coerce', dayfirst=True)

                    if pd.isna(dt_obj) or dt_obj < start_dt or dt_obj > end_dt:
                        print(f"  Skipped holiday at column {col_num} (date: {dt_cell}, parsed: {dt_obj})")
                        skipped_this_yr += 1
                        continue

                    cell = sheet.cell(row=target_row_num, column=col_num)
                    existing_comment = cell.comment
                    existing_style = cell._style if hasattr(cell, '_style') else None
                    cell.value = "f"
                    if existing_comment:
                        cell.comment = existing_comment
                    if existing_style:
                        cell._style = existing_style
                    print(f"  Marked 'f' at cell {cell.coordinate} for {yr}")
                    marked_this_yr += 1
                else:
                    skipped_this_yr += 1
                    print(f"  Skipped non-holiday at column {col_num}")
                    

            #print(f"  {yr}: marked {marked_this_yr}, skipped {skipped_this_yr}")
            total_marked += marked_this_yr
            total_skipped += skipped_this_yr

        #print(f"  Total for {emp_data['full_name']}: {total_marked} marked, {total_skipped} skipped")
        return True
    
    def do_all_holidays(self, out_path=None):
        if not self.emp_list:
            print("No employees yet")
            return None
        
        #print(f"\nProcessing {len(self.emp_list)} people...")
        
        try:
            wb = load_workbook(self.path, data_only=False)
            sheet = wb['IST Stunden']
            # Extract formulas and comments from all sheets
            formulas_comments = {}
            for sheet_name in wb.sheetnames:
                current_sheet = wb[sheet_name]
                for row in current_sheet.iter_rows(min_row=1, max_row=current_sheet.max_row, min_col=1, max_col=current_sheet.max_column):
                    for cell in row:
                        coord = f"{sheet_name}!{cell.coordinate}"
                        formula = cell.value if isinstance(cell.value, str) and cell.value.startswith('=') else None
                        comment = cell.comment.text if cell.comment else None
                        if formula or comment:
                            formulas_comments[coord] = {
                                'formula': formula,
                                'comment': comment,
                                'number_format': cell.number_format if hasattr(cell, 'number_format') else None
                            }
                print(f"Extracted {sum(1 for k in formulas_comments if k.startswith(sheet_name + '!'))} formulas/comments from {sheet_name}")
        except Exception as e:
            #print(f"Can't load file: {e}")
            return None
            
        
        good_count = 0
        for emp in self.emp_list:
            if self.process_employee_holidays(emp, wb,formulas_comments):
                good_count += 1
            else:
                print(f"Failed on {emp['full_name']}")
        
        print(f"\nDone {good_count} of {len(self.emp_list)}")

        sample_verified = 0
        ref_verified = 0
        for emp in self.emp_list[:min(2, len(self.emp_list))]:
            for yr in self.actual_years[:1]:
                if yr in self.yr_offset and emp['state'] in self.bl_mapping.get(yr, {}):
                    target_start = self.yr_offset[yr] + 3
                    target_row_num = target_start + emp['row_offset']
                    date_row = target_start - 2
                    sample_cell = sheet.cell(row=target_row_num, column=self.ist_first)
                    if sample_cell.value == "f":
                        sample_verified += 1
                    print(f"Pre-save check: Cell {sample_cell.coordinate} for {emp['full_name']} ({yr}) has value: {sample_cell.value}")
                    for col in range(self.ist_first, min(self.ist_first + 5, self.ist_last + 1)):
                        formula_cell = sheet.cell(row=date_row, column=col)
                        if isinstance(formula_cell.value, str) and formula_cell.value.startswith('='):
                            print(f"Pre-save check: Found formula {formula_cell.value} at {formula_cell.coordinate}")
                            ref_verified += 1
        for sheet_name in wb.sheetnames:
            if sheet_name != 'IST Stunden':  # Already checked IST Stunden
                other_sheet = wb[sheet_name]
                formula_count = 0
                for row in other_sheet.iter_rows(min_row=1, max_row=10, min_col=1, max_col=10):
                    for cell in row:
                        if isinstance(cell.value, str) and cell.value.startswith('='):
                            formula_count += 1
                            print(f"Pre-save check: Found formula {cell.value} in {sheet_name} at {cell.coordinate}")
                print(f"Pre-save check: Found {formula_count} formulas in {sheet_name} (sample)")
        print(f"Pre-save verification: Found 'f' in {sample_verified} sample cells, valid references in {ref_verified} formula cells")
        # Reapply formulas and comments
        for coord, data in formulas_comments.items():
            sheet_name, cell_coord = coord.split('!')
            target_sheet = wb[sheet_name]
            cell = target_sheet[cell_coord]
            if data['formula']:
                cell.value = data['formula']
            if data['number_format']:
                cell.number_format = data['number_format']
            if data['comment']:
                cell.comment = openpyxl.comments.Comment(data['comment'], 'HolidayTool')
        print(f"Reapplied {len(formulas_comments)} formulas/comments")
        
        if out_path is None:
            out_path = self.path.replace('.xlsx', '_holidays_added.xlsx')
        
        try:
            wb.save(out_path)
            print(f"Saved to: {out_path}")
            wb_verify = load_workbook(out_path, data_only=False)
            sheet_verify = wb_verify['IST Stunden']
            ref_verified_post = 0
            for emp in self.emp_list[:min(2, len(self.emp_list))]:
                for yr in self.actual_years[:1]:
                    if yr in self.yr_offset and emp['state'] in self.bl_mapping.get(yr, {}):
                        target_start = self.yr_offset[yr] + 3
                        date_row = target_start - 2
                        for col in range(self.ist_first, min(self.ist_first + 5, self.ist_last + 1)):
                            formula_cell = sheet_verify.cell(row=date_row, column=col)
                            if isinstance(formula_cell.value, str) and formula_cell.value.startswith('='):
                                print(f"Post-save check: Found formula {formula_cell.value} at {formula_cell.coordinate}")
                                ref_verified_post += 1
            for sheet_name in wb_verify.sheetnames:
                if sheet_name != 'IST Stunden':  # Already checked IST Stunden
                    other_sheet = wb_verify[sheet_name]
                    formula_count = 0
                    for row in other_sheet.iter_rows(min_row=1, max_row=10, min_col=1, max_col=10):
                        for cell in row:
                            if isinstance(cell.value, str) and cell.value.startswith('='):
                                formula_count += 1
                                print(f"Post-save check: Found formula {cell.value} in {sheet_name} at {cell.coordinate}")
                    print(f"Post-save check: Found {formula_count} formulas in {sheet_name} (sample)")
            print(f"Post-save verification: Valid references in {ref_verified_post} formula cells")
            return out_path
        except Exception as e:
            print(f"Save error: {e}")
            return None
    
    def execute(self, out_path=None):
        print("Detecting file format and years...")
        if not self.detect_file_format_and_years():
            print("Could not detect file format")
            return None
        
        
        print("Getting metadata...")
        if not self.get_metadata():
            print("Metadata failed")
            return None
        
        print("\nFinding employees...")
        self.find_employees()
        
        if not self.emp_list:
            print("No employees found")
            return None
        
        print(f"\nGot {len(self.emp_list)} people:")
        for e in self.emp_list:
            print(f"  #{e['emp_num']}: {e['full_name']} ({e['state']})")
        
        print(f"\nMarking holidays...")
        result = self.do_all_holidays(out_path)
        
        if result:
            print(f"\nDone! Processed {len(self.emp_list)} employees")
            print(f"File format: {self.file_format} employees")
            print(f"Years processed: {self.actual_years}")
            print(f"File: {result}")
        else:
            print("\nFailed")
        
        return result
    
    def change_max(self, new_max):
        if new_max < 1 or new_max > 50:
            print(f"Max should be 1-50, got {new_max}")
            return False
        
        self.max_emp = new_max
        print(f"Max set to: {new_max}")
        return True

app = Flask(__name__)
CORS(app)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
temp_files = {}

@app.route('/api/health')
def health_check():
    return jsonify({'status': 'ok', 'message': 'Backend is running'})

@app.route('/api/process-holidays', methods=['POST'])
def process_holidays():
    try:
        holiday_lines = [
            "Spreading holidays like cheese on a hot pizza 🍕📅",
            "Distributing holidays like Nutella on warm toast 🧈📆",
            "Layering holidays like frosting on a cake 🎂📅",
            "Smearing holidays across your calendar like butter on bread 🧈🗓️",
            "Sprinkling holidays like herbs on a Margherita 🍃🍕"
        ]
        fun_message = random.choice(holiday_lines)

        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.endswith('.xlsx'):
            return jsonify({'error': 'Currently only accepting Excel Files'}), 400
        
      
        temp_dir = tempfile.mkdtemp()
        input_file = os.path.join(temp_dir, secure_filename(file.filename))
        file.save(input_file)
        
   
        max_employees = request.form.get('max_employees', 30, type=int)
        
     
        tool = HolidayTool(input_file)
        tool.change_max(max_employees)
        
       
        output_filename = f"processed_{uuid.uuid4().hex}.xlsx"
        output_file = os.path.join(temp_dir, output_filename)
        
        result = tool.execute(output_file)
        
        if result:
  
            file_id = str(uuid.uuid4())
            temp_files[file_id] = {
                'path': result,
                'filename': output_filename,
                'temp_dir': temp_dir
            }
            
            return jsonify({
                'success': True,
                'message': fun_message,
                'employees_processed': len(tool.emp_list),
                'file_format_detected': f"{tool.file_format} employees",
                'years_processed': tool.actual_years,
                'download_url': f'/api/download/{file_id}'
            })
        else:
           
            shutil.rmtree(temp_dir, ignore_errors=True)
            return jsonify({'error': 'Processing failed'}), 500
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/download/<file_id>')
def download_file(file_id):
    if file_id not in temp_files:
        return jsonify({'error': 'File not found'}), 404
    
    file_info = temp_files[file_id]
    try:
        return send_file(
            file_info['path'], 
            as_attachment=True, 
            download_name=file_info['filename']
        )
    except Exception as e:
        return jsonify({'error': 'Download failed'}), 500

if __name__ == "__main__":
    app.run(debug=True, host='0.0.0.0', port=5000)
